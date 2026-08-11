import re
import requests
import io
import zipfile
from urllib.parse import quote
from openpyxl import load_workbook
from auth import auth_headers
from config import SHEET_ID, DAY_SHEETS, HOMEWORK_SHEETS

SHEETS_API = "https://sheets.googleapis.com/v4/spreadsheets"

_xlsx_cache: bytes | None = None
_sheet_data_cache: dict[str, list[list[str]]] = {}


def match_full_name(short_name: str, full_names: list[str]) -> str | None:
    """짧은 이름(서영)으로 전체 이름(이서영) 찾기"""
    for full in full_names:
        if full.endswith(short_name) or short_name in full:
            return full
    return None


def _get_xlsx(force_refresh: bool = False) -> bytes:
    global _xlsx_cache
    if _xlsx_cache is None or force_refresh:
        if force_refresh:
            _sheet_data_cache.clear()
        url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
        resp = requests.get(url, headers=auth_headers(), allow_redirects=True)
        resp.raise_for_status()
        _xlsx_cache = resp.content
    return _xlsx_cache


def clear_cache():
    global _xlsx_cache
    _xlsx_cache = None
    _sheet_data_cache.clear()


def _fetch_sheet_data(sheet_name: str) -> list[list[str]]:
    if sheet_name in _sheet_data_cache:
        return _sheet_data_cache[sheet_name]
    xlsx_bytes = _get_xlsx()
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    _sheet_data_cache[sheet_name] = rows
    return rows


def get_student_groups(day: str) -> list[dict]:
    """
    시간표 탭에서 시간대별 학생 그룹을 추출.
    반환: [{"time": "14:35", "students": ["배서준", ...]}, ...]
    """
    sheet_name = DAY_SHEETS.get(day)
    if not sheet_name:
        return []

    rows = _fetch_sheet_data(sheet_name)

    # 불필요한 키워드 (학생 이름이 아닌 것)
    NON_STUDENT = {"부스", "후방데스크", "플루언트", ""}

    def clean_names(bracket_text: str) -> list[str]:
        # 주석 및 부가설명 제거
        text = re.sub(r"\*[^,<>]*", "", bracket_text)   # *재희 6월 신규
        text = re.sub(r'"[^"]*"', "", text)              # "플루언트"
        text = re.sub(r'\([^)]*\)', "", text)            # (플루언트)
        text = re.sub(r'-\s*고등', "", text)             # -고등
        text = re.sub(r'/플루언트', "", text)            # /플루언트
        text = re.sub(r':\s*플루언트', "", text)         # :플루언트
        # 쉼표 또는 공백으로 분리
        parts = re.split(r'[,\s]+', text)
        names = []
        for name in parts:
            name = name.strip().strip("()")
            if name and name not in NON_STUDENT and not re.match(r"^\d", name) and len(name) >= 2:
                names.append(name)
        return names

    current_time = None
    groups = []

    for row in rows:
        time_cell = row[0].strip().split("\n")[0] if row else ""
        time_match = re.match(r"(\d{1,2}:\d{2})", time_cell)
        if time_match:
            current_time = time_match.group(1)

        # 학생 이름 추출: <이름,...> 패턴 (부스/후방 제외)
        all_text = " ".join(row)
        brackets = re.findall(r"<([^>]+)>", all_text)

        for bracket in brackets:
            if any(kw in bracket for kw in ["부스", "후방", "1,2강"]):
                continue
            names = clean_names(bracket)
            if not names:
                continue
            # 같은 시간대에 이미 같은 그룹이 있으면 추가 안 함
            existing = next((g for g in groups if g["time"] == current_time and set(g["students"]) == set(names)), None)
            if not existing and current_time:
                groups.append({"time": current_time, "students": names})

    return groups


def get_student_progress(day: str) -> dict[str, dict]:
    """
    해당 요일 과제 탭에서 학생별 진도 정보 추출.
    반환: {"학생명": {"lesson": 8, "level": 1, "grammar_ref": "B4-3"}, ...}
    """
    sheet_name = HOMEWORK_SHEETS.get(day)
    rows = _fetch_sheet_data(sheet_name) if sheet_name else []
    progress = {}

    for row in rows:
        if len(row) < 3:
            continue
        name = row[0].strip()
        boothwork = row[1].strip() if len(row) > 1 else ""
        homework = row[2].strip() if len(row) > 2 else ""

        if not name or name in ("이름", ""):
            continue

        # B열(Boothwork), C열(Homework)에서만 소스 감지 (피드백 칸 오인식 방지)
        progress_text = f"{boothwork} {homework}"
        full_row_text = " ".join(row)
        is_fluent = "플루언트" in progress_text or "fluent" in progress_text.lower()
        lesson_num = None
        unit_num = None
        level_num = None

        lv_dash = re.search(r"[Ll][Vv](\d+)-(\d+)과?", full_row_text)
        if lv_dash:
            level_num = int(lv_dash.group(1))
            if is_fluent:
                unit_num = int(lv_dash.group(2))
            else:
                lesson_num = int(lv_dash.group(2))
        else:
            level_match = re.search(r"[Ll]evel\s*(\d+)|LV\s*(\d+)|Lv\s*(\d+)|레벨\s*(\d+)", full_row_text)
            if level_match:
                level_num = int(next(g for g in level_match.groups() if g))

            if is_fluent:
                unit_match = re.search(r"[Uu]nit\s*(\d+)|유닛\s*(\d+)", full_row_text)
                if unit_match:
                    unit_num = int(next(g for g in unit_match.groups() if g))
            else:
                lesson_match = re.search(r"[Ll]esson\s*(\d+)|레슨\s*(\d+)", full_row_text)
                if lesson_match:
                    lesson_num = int(next(g for g in lesson_match.groups() if g))

        # B열(Boothwork): 문법 참조 추출 (ITC 전용, 예: B4-3, P1-2)
        grammar_ref = None
        if not is_fluent:
            grammar_match = re.search(r"문법\s*([BPIAbpia])(\d+)-?(\d*)", boothwork)
            if grammar_match:
                prefix = grammar_match.group(1).upper()
                unit = grammar_match.group(2)
                sub = grammar_match.group(3)
                grammar_ref = f"{prefix}{unit}-{sub}" if sub else f"{prefix}{unit}"

        progress[name] = {
            "source": "fluent" if is_fluent else "itc",
            "level": level_num,
            "lesson": lesson_num,       # ITC 전용
            "unit": unit_num,           # 플루언트 전용
            "grammar_ref": grammar_ref, # ITC 전용
            "homework_raw": homework,
            "boothwork_raw": boothwork,
        }

    return progress


def get_student_checks(day: str, student_names: list[str]) -> dict[str, dict]:
    """
    해당 요일 과제 탭에서 특정 학생들의 체크 항목을 Sheets API로 '실시간' 읽기.
    ※ xlsx export 는 몇십 초 지연이 있어 방금 시트에서 바꾼 값이 안 보일 수 있으므로
      values API 로 직접 조회한다(지연 없음).
    반환: {"학생명": {"과제완수": bool, "레슨통과": bool, "부스클리닉": bool}, ...}
    D열=지각1회, E열=지각2회, F열=과제완수, G열=레슨통과, H열=프리토킹, I열=부스&클리닉
    """
    sheet_name = HOMEWORK_SHEETS.get(day)
    if not sheet_name:
        return {}

    rng = quote(f"'{sheet_name}'!A:I", safe="")
    url = f"{SHEETS_API}/{SHEET_ID}/values/{rng}?valueRenderOption=UNFORMATTED_VALUE"
    resp = requests.get(url, headers=auth_headers())
    resp.raise_for_status()
    values = resp.json().get("values", [])

    def _is_true(v) -> bool:
        # 체크박스는 UNFORMATTED_VALUE 로 읽으면 파이썬 bool True/False 로 온다.
        return v is True or str(v).strip().upper() == "TRUE"

    result = {}
    for row in values:
        if not row or not str(row[0]).strip():
            continue
        name = str(row[0]).strip()
        if name not in student_names:
            continue
        result[name] = {
            "과제완수":  _is_true(row[5]) if len(row) > 5 else False,   # F열
            "레슨통과":  _is_true(row[6]) if len(row) > 6 else False,   # G열
            "부스클리닉": _is_true(row[8]) if len(row) > 8 else False,   # I열
        }
    return result


def _find_student_row(sheet_name: str, student_name: str) -> int | None:
    """시트에서 학생 행 번호 반환 (1-based). 없으면 None."""
    rows = _fetch_sheet_data(sheet_name)
    for idx, row in enumerate(rows):
        if row and row[0].strip() == student_name:
            return idx + 1
    return None


def _write_cell(sheet_name: str, row_num: int, col_letter: str, value) -> bool:
    """시트의 특정 셀에 값 쓰기."""
    range_notation = f"'{sheet_name}'!{col_letter}{row_num}"
    url = f"{SHEETS_API}/{SHEET_ID}/values/{quote(range_notation, safe='')}?valueInputOption=RAW"
    resp = requests.put(
        url,
        headers={**auth_headers(), "Content-Type": "application/json"},
        json={"values": [[value]]}
    )
    return resp.status_code == 200


def write_lesson_check(day: str, student_name: str, field: str, value: bool) -> bool:
    """
    레슨통과(G열) 또는 부스클리닉(I열) 체크값을 시트에 기록.
    field: "레슨통과" | "부스클리닉"
    """
    sheet_name = HOMEWORK_SHEETS.get(day)
    if not sheet_name:
        return False
    col = "G" if field == "레슨통과" else "I"
    row_num = _find_student_row(sheet_name, student_name)
    if not row_num:
        return False
    return _write_cell(sheet_name, row_num, col, value)


def advance_student_lesson(day: str, student_name: str, is_fluent: bool) -> bool:
    """
    시트 B열(Boothwork)과 C열(Homework)에서 레슨/유닛 번호를 +1로 업데이트.
    """
    sheet_name = HOMEWORK_SHEETS.get(day)
    if not sheet_name:
        return False
    row_num = _find_student_row(sheet_name, student_name)
    if not row_num:
        return False

    rows = _fetch_sheet_data(sheet_name)
    row = rows[row_num - 1]
    boothwork = row[1] if len(row) > 1 else ""
    homework  = row[2] if len(row) > 2 else ""

    def increment_num(text: str, pattern: str) -> str:
        def replacer(m):
            return m.group(1) + str(int(m.group(2)) + 1)
        return re.sub(pattern, replacer, text)

    if is_fluent:
        new_bw = increment_num(boothwork, r'([Uu]nit\s*)(\d+)')
        new_hw = increment_num(homework,  r'([Uu]nit\s*)(\d+)')
    else:
        new_bw = increment_num(boothwork, r'([Ll]esson\s*)(\d+)')
        new_hw = increment_num(homework,  r'([Ll]esson\s*)(\d+)')

    ok1 = _write_cell(sheet_name, row_num, "B", new_bw)
    ok2 = _write_cell(sheet_name, row_num, "C", new_hw)
    return ok1 and ok2


def write_freetalk_points(day: str, student_name: str, points: int) -> bool:
    """
    해당 요일 과제 탭의 H열(프리토킹)에 점수 기록.
    학생 이름으로 행을 찾아 H열에 숫자 덮어쓰기.
    """
    sheet_name = HOMEWORK_SHEETS.get(day)
    if not sheet_name:
        return False

    # 행 번호 찾기 (1-based)
    rows = _fetch_sheet_data(sheet_name)
    row_num = None
    for idx, row in enumerate(rows):
        if row and row[0].strip() == student_name:
            row_num = idx + 1  # 1-based
            break

    if not row_num:
        return False

    # H열 = 8번째 열
    range_notation = f"'{sheet_name}'!H{row_num}"
    url = f"{SHEETS_API}/{SHEET_ID}/values/{quote(range_notation, safe='')}?valueInputOption=RAW"
    resp = requests.put(
        url,
        headers={**auth_headers(), "Content-Type": "application/json"},
        json={"values": [[points]]}
    )
    return resp.status_code == 200
